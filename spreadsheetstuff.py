import openpyxl
from openpyxl.workbook import workbook
from openpyxl import load_workbook

def remove_cols(ws, start, width, msg):
    ws.delete_cols(start, width)
    print(f'\n{}msg}\n{"-" * 20}')

def rename_headings(ws, cell, value):
    ws[cell].value = value

wb = load_workbook('RCW.xlsx')
sheet = wb.sheetnames # adds sheet names to variable sheet
ws = wb[sheet[0]]

    col_dels = ((1, 4, 'Columns A-D Deleted'),
                (11, 14, 'Columns K-X Deleted'),
                (12, 4, 'Columns L-O Deleted'),
                (14, 3, 'Columns N-P Deleted'),
                (15, 2, 'Columns O-P Deleted'),
                (16, 1, 'Column P Deleted'),
                (18, 18, 'Columns R-AI Deleted'),
                (19, 5, 'Columns S-W Deleted')
                )

    col_heads = (('N1', 'Contract End'),
                 ('P1', 'Has FTTP'),
                 ('Q1', 'Greenfield'),
                 ('R1', 'TV Customer')
                 )



    print('Removing columns')
    for start, width, msg in col_dels:
        remove_cols(ws, start, width, msg)

    print('Renaming Headings')
    for cell, msg in col_heads:
        rename_headings(ws, cell, msg)
print('--------------------------------------------------')
print('Moving Columns')
print('--------------------------------------------------')
ws.insert_cols(1, amount=1)

col_p = ws['P']
for idx, cell in enumerate(col_p,1):
    ws.cell(row = idx, column = 1).value = cell.value

ws.delete_cols(16, 1)

ws.insert_cols(15, amount=2)

col_fttp = ws['R']
for idx, cell in enumerate(col_fttp,1):
    ws.cell(row = idx, column = 15).value = cell.value

col_greenfield = ws['S']
for idx, cell in enumerate(col_greenfield, 1):
    ws.cell(row=idx, column=16).value = cell.value
print('Cleaning Up')
ws.delete_cols(18, 2)
ws.delete_cols(19, 1)
print('--------------------------------------------------')
num_sheets = int(input("How many Sheets: "))
total_rows = ws.max_row-1 # counts the total rows with data
total = int(total_rows // num_sheets)
total_cols = ws.max_column
print('--------------------------------------------------')
#wb.save("RCW.xlsx")
print('Data Completed')
print('--------------------------------------------------')
print(f"You Have a Total of {total_rows} records\n")

print(f"You have {total} of records for each day")
print('--------------------------------------------------')
print('Splitting Data')


sheet_count = 0

while sheet_count < num_sheets:
        new_sheet = wb.create_sheet("Sheet_"+ str(x))
        sheet_count = sheet_count + 1
        total_to_copy = total
        row_offset = 0
        for i in range(1, total_to_copy):
            for j in range(1, ws.max_column):
                new_sheet.cell(row=i, column=j).value = ws.cell(row=i, column=j).value
                row_offset = total_to_copy + 1


wb.save("RCW.xlsx")